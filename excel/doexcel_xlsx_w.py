#!/usr/bin/env python
#!coding=utf-8
#author=guolin
from openpyxl import Workbook
import sys
reload(sys)
sys.setdefaultencoding( "utf-8")

file_path = r"./test.xlsx";

def writefile():
    print u'开始处理excel'
    # 创建文档对象
    wb = Workbook()

    # 创建sheet表，表名称，表位置index=0表示第一个位置
    sheet = wb.create_sheet('mbean_data', index=0)

    # 添加标题
    sheet.append(["姓名" ,"年龄", "性别", "专业"])

    # 添加数据
    sheet.append(["张三" ,18, "女", "html"])
    sheet.append(["李四" ,25, "男", "java"])

    # 设置第1行行高
    sheet.row_dimensions[1].height = 50
    # 设置A列列宽
    sheet.column_dimensions['A'].width = 20

    # 设置A6的位置为B2+B3求和
    sheet['A6'] = '=SUM(B2:B3)'

    # 获得最大列和最大行
    print u'当前excel的行数=',sheet.max_row
    print u'当前excel的列数=',sheet.max_column

    # 文件写入路径
    wb.save(file_path)

    print u'excel保存完成,文件地址=',file_path

if __name__=="__main__":
    writefile()