#!/usr/bin/env python
#!coding=utf-8
#author=guolin
from openpyxl import load_workbook
import sys
reload(sys)
sys.setdefaultencoding( "utf-8")

file_path = r"./test.xlsx";

def readfile():
    print u'开始处理excel'
    # 读取文档对象
    wb = load_workbook(file_path)

    # 获得所有sheet的名称
    print u'sheet列表=',wb.sheetnames
    # 根据sheet名字获得sheet
    sheet = wb['mbean_data']
    # 获得sheet名
    print u'当前sheet表名称=',sheet.title

    # 获取指定位置的值
    print u'当前sheet表的A1位置的值=',sheet['A1'].value

    # 除了用下标的方式获得，还可以用cell函数, 换成数字，这个表示B4
    print u'根据行和列获取指定的值=', sheet.cell(row=1, column=2).value


    # 按照行循环,从0开始计算
    print u'按照行循环'
    for row in sheet.rows:
        print row[0].value
    # 按照列循环,从0开始计算
    print u'按照列循环'
    for column in sheet.columns:
        print column[0].value


if __name__=="__main__":
    readfile()